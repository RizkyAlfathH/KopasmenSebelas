from time import timezone
from urllib import request
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from admin_koperasi.models import Admin
from .models import Anggota
from .forms import AdminForm, AnggotaForm
from django.db import connection
from django.db import transaction
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from .models import Anggota
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from .models import Anggota
from django.shortcuts import render, redirect
from django.db.models import Sum
from admin_koperasi.models import Admin
from .models import Anggota
from django.contrib import messages
from django.http import JsonResponse
import pandas as pd
from django.contrib import messages
from simpanan.models import Simpanan, JenisSimpanan
from django.contrib.auth.hashers import make_password
from decimal import Decimal, InvalidOperation
import re
from datetime import datetime, date
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook


def kelola_akun(request):
    # Cek login session
    if not request.session.get('admin_id'):
        return redirect('admin_koperasi:login')

    role = request.session.get('admin_role')
    username = request.session.get('admin_username')

    # Ambil filter pencarian
    search_admin = request.GET.get("searchAdmin", "")
    search_anggota = request.GET.get("searchAnggota", "")

    # === ADMIN ===
    admins = Admin.objects.all()
    if search_admin:
        admins = admins.filter(username__icontains=search_admin)
    admins = admins.order_by("id_admin")

    paginator_admin = Paginator(admins, 10)
    page_admin = request.GET.get("page_admin", 1)
    admins_page = paginator_admin.get_page(page_admin)

    # === ANGGOTA ===
    anggotas = Anggota.objects.all()
    if search_anggota:
        anggotas = anggotas.filter(nama__icontains=search_anggota)
    anggotas = anggotas.order_by("nomor_anggota")

    paginator_anggota = Paginator(anggotas, 10)
    page_anggota = request.GET.get("page_anggota", 1)
    anggotas_page = paginator_anggota.get_page(page_anggota)

    context = {
        "username": username,
        "role": role,
        "admins": admins_page,
        "anggotas": anggotas_page,
        "searchAdmin": search_admin,
        "searchAnggota": search_anggota,
    }
    return render(request, "kelola_akun.html", context)

# ===============================
# TAMBAH / EDIT / HAPUS ADMIN
# ===============================
def tambah_admin(request):
    if not request.session.get('admin_id'):
        return redirect('admin_koperasi:login')
    role = request.session.get('admin_role')
    username = request.session.get('admin_username')

    if request.method == 'POST':
        form = AdminForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Admin berhasil ditambahkan.")
            return redirect('kelola_akun')
    else:
        form = AdminForm()

    return render(request, 'form_admin.html', {'form': form, 'judul': 'Tambah Admin', 'username': username, 'role': role})

def edit_admin(request, id_admin):
    if not request.session.get('admin_id'):
        return redirect('admin_koperasi:login')
    role = request.session.get('admin_role')
    username = request.session.get('admin_username')

    admin = get_object_or_404(Admin, id_admin=id_admin)
    form = AdminForm(request.POST or None, instance=admin)
    if form.is_valid():
        form.save()
        messages.success(request, "Admin berhasil diperbarui.")
        return redirect('kelola_akun')

    return render(request, 'form_admin.html', {'form': form, 'judul': 'Edit Admin', 'username': username, 'role': role})

def hapus_admin(request, id_admin):
    admin = get_object_or_404(Admin, id_admin=id_admin)
    admin.delete()
    messages.success(request, "Admin berhasil dihapus.")
    return redirect('kelola_akun')

def detail_admin(request, id_admin):
    if not request.session.get('admin_id'):
        return redirect('admin_koperasi:login')
    role = request.session.get('admin_role')
    username = request.session.get('admin_username')

    admin = get_object_or_404(Admin, id_admin=id_admin)
    all_admins = Admin.objects.order_by('id_admin')
    nomor_urut = list(all_admins).index(admin) + 1
    return render(request, 'detailA.html', {'admin': admin, 'nomor_urut': nomor_urut, 'username': username, 'role': role})

# ===============================
# TAMBAH / EDIT / HAPUS ANGGOTA
# ===============================
def tambah_anggota(request):
    if not request.session.get('admin_id'):
        return redirect('admin_koperasi:login')
    role = request.session.get('admin_role')
    username = request.session.get('admin_username')

    if request.method == 'POST':
        form = AnggotaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Anggota berhasil ditambahkan.")
            return redirect('kelola_akun')
    else:
        form = AnggotaForm()

    return render(request, 'form_anggota.html', {'form': form, 'judul': 'Tambah Anggota', 'username': username, 'role': role})

def edit_anggota(request, nomor_anggota):
    if not request.session.get('admin_id'):
        return redirect('admin_koperasi:login')
    role = request.session.get('admin_role')
    username = request.session.get('admin_username')

    anggota = get_object_or_404(Anggota, nomor_anggota=nomor_anggota)
    form = AnggotaForm(request.POST or None, instance=anggota)
    if form.is_valid():
        form.save()
        messages.success(request, "Anggota berhasil diperbarui.")
        return redirect('kelola_akun')

    return render(request, 'form_anggota.html', {'form': form, 'judul': 'Edit Anggota', 'username': username, 'role': role})

def hapus_anggota(request, nomor_anggota):
    anggota = get_object_or_404(Anggota, nomor_anggota=nomor_anggota)
    anggota.delete()
    messages.success(request, "Anggota berhasil dihapus.")
    return redirect('kelola_akun')

def detail_anggota(request, nomor_anggota):
    if not request.session.get('admin_id'):
        return redirect('admin_koperasi:login')
    role = request.session.get('admin_role')
    username = request.session.get('admin_username')

    anggota = get_object_or_404(Anggota, nomor_anggota=nomor_anggota)
    return render(request, 'detail.html', {'anggota': anggota, 'username': username, 'role': role})


def export_excel_anggota(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daftar Anggota"


    headers = ["No. Anggota", "Nama", "NIP", "Alamat", "No. Telepon", "Email", "Jenis Kelamin", "Tanggal Daftar", "Status", "Tanggal Nonaktif", "Alasan Nonaktif" ]
    ws.append(headers)


    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # kuning
    header_font = Font(bold=True, color="000000")  # teks hitam tebal
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border


    for anggota in Anggota.objects.all():
        ws.append([
            anggota.nomor_anggota,
            anggota.nama,
            anggota.nip,
            anggota.alamat,
            anggota.no_telp,
            anggota.email,
            anggota.jenis_kelamin,
            anggota.tanggal_daftar.strftime("%d-%m-%Y"),
            anggota.status,
            anggota.tanggal_nonaktif,
            anggota.alasan_nonaktif
        ])

   
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")

    
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Response
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="anggota koperasi.xlsx"'
    wb.save(response)
    return response

from reportlab.lib.pagesizes import landscape, A4
def export_pdf_anggota(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="anggota koperasi.pdf"'

    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                        leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
    elements = []

    styles = getSampleStyleSheet()
    normal_style = styles["Normal"]  
    normal_style.fontSize = 7

    title = Paragraph("Daftar Anggota", styles['Title'])
    elements.append(title)

    
    data = [["No. Anggota", "Nama", "NIP", "Alamat", "No. Telepon", 
             "Email", "Jenis Kelamin", "Tanggal Daftar", "Status","Tanggal Nonaktif", "Alasan Nonaktif" ]]

    
    for anggota in Anggota.objects.all():
        data.append([
            Paragraph(str(anggota.nomor_anggota or ""), normal_style),
            Paragraph(anggota.nama or "", normal_style),
            Paragraph(anggota.nip or "", normal_style),
            Paragraph(anggota.alamat or "", normal_style),
            Paragraph(anggota.no_telp or "", normal_style),
            Paragraph(anggota.email or "", normal_style),
            Paragraph(anggota.jenis_kelamin or "", normal_style),
            Paragraph(
                anggota.tanggal_daftar.strftime("%d-%m-%Y") if anggota.tanggal_daftar else "",
                normal_style
            ),
            Paragraph(anggota.status or "", normal_style),
            Paragraph(
                anggota.tanggal_nonaktif.strftime("%d-%m-%Y") if anggota.tanggal_nonaktif else "",
                normal_style
            ),
            Paragraph(anggota.alasan_nonaktif or "", normal_style)
        ])

    col_widths = [45, 80, 55, 110, 65, 80, 50, 60, 45, 65, 75]


    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("FONTSIZE", (0,0), (-1,-1), 7),
        ("BACKGROUND", (0,0), (-1,0), colors.yellow),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("ALIGN", (0,0), (-1,0), "CENTER"),  
        ("ALIGN", (0,1), (-1,-1), "LEFT"),   
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,0), 5),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
    ]))

    elements.append(table)
    doc.build(elements)
    return response

# API untuk cek email unik (dipakai AJAX di JS)
def cek_email(request):
    email = request.GET.get("email", "")
    exists = Anggota.objects.filter(email=email).exists()
    return JsonResponse({"exists": exists})

def baca_data_anggota(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active  # ambil sheet pertama

    data = []
    for row in sheet.iter_rows(min_row=5):  # mulai dari baris ke-5 (karena header di atas)
        nomor_anggota = row[1].value or "-"
        nama = row[2].value or "-"
        umur = row[3].value or "-"
        jenis_kelamin = row[4].value or "-"
        pekerjaan = row[5].value or "-"
        alamat = row[6].value or "-"
        tanggal_daftar = row[7].value or "-"
        alasan_nonaktif = row[13].value or "-"
        tanggal_nonaktif = row[12].value or "-"

        data.append({
            "nomor_anggota": nomor_anggota,
            "nama": nama,
            "umur": umur,
            "jenis_kelamin": jenis_kelamin,
            "pekerjaan": pekerjaan,
            "alamat": alamat,
            "tanggal_daftar": tanggal_daftar,
            "alasan_nonaktif": alasan_nonaktif,
            "tanggal_nonaktif": tanggal_nonaktif,
        })
    
    return data

def upload_excel(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        try:
            file = request.FILES["excel_file"]
            wb = load_workbook(file, data_only=True)
            ws = wb.active

            data = []
            start_row = 5

            for row in ws.iter_rows(min_row=start_row):
                nomor_anggota = row[1].value or "-"
                nama = row[2].value or "-"
                umur = row[3].value or None
                jenis_kelamin = row[4].value or "-"
                pekerjaan = row[5].value or "-"
                alamat = row[6].value or "-"
                tanggal_masuk = row[7].value or None
                tanggal_berhenti = row[12].value or None
                alasan_nonaktif = row[13].value or "-"

                # Konversi jenis kelamin
                if isinstance(jenis_kelamin, str):
                    jk = jenis_kelamin.strip().upper()
                    jenis_kelamin = "Laki-laki" if jk == "L" else "Perempuan" if jk == "P" else "-"

                # Konversi tanggal ke format date
                tgl_masuk = tanggal_masuk.date() if isinstance(tanggal_masuk, datetime) else None
                tgl_nonaktif = tanggal_berhenti.date() if isinstance(tanggal_berhenti, datetime) else None

                # Lewati baris kosong
                if nama == "-" and nomor_anggota == "-":
                    continue

                anggota, created = Anggota.objects.update_or_create(
                    nomor_anggota=nomor_anggota,
                    defaults={
                        "nama": nama,
                        "umur": umur,
                        "jenis_kelamin": jenis_kelamin,
                        "pekerjaan": pekerjaan,
                        "alamat": alamat,
                        "tanggal_daftar": tgl_masuk  or date.today(),  # ✅ fix
                        "tanggal_nonaktif": tgl_nonaktif,
                        "alasan_nonaktif": alasan_nonaktif,
                        "status": "nonaktif" if tgl_nonaktif else "aktif",
                    }
                )

                # Kalau belum ada password, isi default
                if not anggota.password_hash:
                    anggota.set_password("12345")
                    anggota.save()

                data.append({
                    "nomor_anggota": anggota.nomor_anggota,
                    "nama": anggota.nama,
                    "status": "Ditambahkan" if created else "Diperbarui"
                })

            return render(request, "kelola_akun.html", {
                "success": True,
                "total_data": len(data),
                "data": data,
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return render(request, "kelola_akun.html", {
                "success": False,
                "message": f"❌ Kesalahan saat membaca atau menyimpan data: {str(e)}"
            })

    return render(request, "kelola_akun.html", {
        "success": False,
        "message": "Tidak ada file Excel dikirim."
    })


# def _to_decimal(val):
#     if pd.isna(val):
#         return Decimal(0)
#     s = str(val).strip()
#     if s in ("", "-", "—"):
#         return Decimal(0)
#     s2 = re.sub(r"[^\d\.\-]", "", s)
#     if s2 in ("", ".", "-"):
#         return Decimal(0)
#     try:
#         return Decimal(s2)
#     except InvalidOperation:
#         s3 = s2.replace(".", "")
#         if s3 == "":
#             return Decimal(0)
#         return Decimal(s3)


# # helper parse metadata (nama, nomor anggota, nip)
# def _parse_metadata_from_df(df):
#     for i in range(min(6, len(df))):
#         rowvals = [str(x) for x in df.iloc[i].dropna().astype(str).tolist()]
#         if not rowvals:
#             continue
#         joined = " ".join(rowvals)
#         jup = joined.upper()
#         if "NAMA" in jup and ("NO" in jup and "ANGGOTA" in jup or "NO. ANGGOTA" in jup):
#             raw = joined
#             m_nama = re.search(r"NAMA\s*:?\s*(.*?)\s+(?:NIP|NO\.?\s?ANGGOTA|HALAMAN|$)", raw, flags=re.IGNORECASE)
#             nama = (m_nama.group(1).strip() if m_nama else None)
#             m_nip = re.search(r"NIP\s*:?\s*([A-Za-z0-9\-/]+)", raw, flags=re.IGNORECASE)
#             nip = (m_nip.group(1).strip() if m_nip else None)
#             m_no = re.search(r"(?:NO\.?\s?ANGGOTA|NO ANGGOTA)\s*:?\s*([A-Za-z0-9\-]+)", raw, flags=re.IGNORECASE)
#             no = (m_no.group(1).strip() if m_no else None)
#             return {"nomor_anggota": no, "nama": nama, "nip": nip, "meta_row_idx": i}
#     return None


# # helper cari baris header transaksi
# def _find_header_row_index(df, start_row=0, lookahead=12):
#     max_row = min(len(df), start_row + lookahead)
#     for i in range(start_row, max_row):
#         row_texts = [str(x).upper() if not pd.isna(x) else "" for x in df.iloc[i].tolist()]
#         joined = " ".join(row_texts)
#         if "POKOK" in joined or "WAJIB" in joined or "SUKARELA" in joined or "TANGGAL" in joined:
#             return i
#     return None


# def safe_decimal(value):
#     """Konversi aman dari nilai Excel ke Decimal"""
#     if value is None:
#         return Decimal("0")
#     if isinstance(value, (int, float, Decimal)):
#         return Decimal(str(value))
#     try:
#         s = str(value).strip()
#         s = re.sub(r"[^\d,\.\-]", "", s)  # hapus simbol non-angka
#         if s == "" or s == "-":
#             return Decimal("0")
#         s = s.replace(",", ".")
#         return Decimal(s)
#     except (InvalidOperation, ValueError):
#         return Decimal("0")


# def upload_excel(request):
#     """
#     Import data simpanan anggota dari file Excel (multi-sheet: tiap tahun)
#     Baris 4 = total saldo sebelumnya
#     Baris 5,8,11,dst = transaksi bulanan
#     """
#     if request.method == "POST" and request.FILES.get("excel_file"):
#         try:
#             file = request.FILES["excel_file"]
#             wb = load_workbook(file, data_only=True)

#             total_import = 0
#             nama = nomor_anggota = None
#             admin = Admin.objects.first()

#             for sheet_name in wb.sheetnames:
#                 ws = wb[sheet_name]
#                 print(f"🔹 Membaca sheet: {sheet_name}")

#                 # Skip otomatis jika bukan sheet transaksi (cek apakah ada tanggal di B5)
#                 if not ws["B5"].value and not ws["D4"].value:
#                     print(f"   ⚠️ Sheet {sheet_name} dilewati (tidak ada data transaksi).")
#                     continue

#                 if not nama:
#                     nama = ws["D1"].value or "-"
#                 if not nomor_anggota:
#                     nomor_anggota = ws["K1"].value or "-"

#                 nip = f"NA-{nomor_anggota}"
#                 anggota_obj, _ = Anggota.objects.update_or_create(
#                     nomor_anggota=nomor_anggota,
#                     defaults={
#                         "nama": nama,
#                         "nip": nip,
#                         "alamat": "-",
#                         "no_telp": "-",
#                         "email": "-",
#                         "jenis_kelamin": "-",
#                         "status": "-",
#                     },
#                 )

#                 # ✅ Gunakan safe_decimal agar aman dari teks atau simbol
#                 prev_total_pokok = safe_decimal(ws["D4"].value)
#                 prev_total_wajib = safe_decimal(ws["E4"].value)
#                 prev_total_sukarela = safe_decimal(ws["F4"].value)
#                 prev_total_danasos = safe_decimal(ws["O4"].value)

#                 print(
#                     f"   → Saldo awal: Pokok={prev_total_pokok}, "
#                     f"Wajib={prev_total_wajib}, Sukarela={prev_total_sukarela}, Dana Sosial={prev_total_danasos}"
#                 )

#                 pokok_jenis = JenisSimpanan.objects.get(nama_jenis="Simpanan Pokok")
#                 wajib_jenis = JenisSimpanan.objects.get(nama_jenis="Simpanan Wajib")
#                 sukarela_jenis = JenisSimpanan.objects.get(nama_jenis="Simpanan Sukarela")

#                 start_row = 5
#                 max_row = ws.max_row

#                 while start_row <= max_row:
#                     tanggal = ws[f"B{start_row}"].value
#                     if not tanggal:
#                         start_row += 3
#                         continue

#                     # ✅ Parsing tanggal aman
#                     if isinstance(tanggal, datetime):
#                         tanggal_fix = tanggal.date()
#                     elif isinstance(tanggal, date):
#                         tanggal_fix = tanggal
#                     elif isinstance(tanggal, str):
#                         try:
#                             tanggal_fix = datetime.strptime(tanggal.strip(), "%d-%b-%y").date()
#                         except Exception:
#                             try:
#                                 tanggal_fix = datetime.strptime(tanggal.strip(), "%Y-%m-%d").date()
#                             except Exception:
#                                 tanggal_fix = datetime.now().date()
#                     else:
#                         tanggal_fix = datetime.now().date()

#                     # ✅ Konversi nilai simpanan
#                     pokok = safe_decimal(ws[f"D{start_row}"].value)
#                     wajib = safe_decimal(ws[f"E{start_row}"].value)
#                     sukarela = safe_decimal(ws[f"F{start_row}"].value)
#                     dana_sosial = safe_decimal(ws[f"O{start_row}"].value)

#                     if pokok > 0:
#                         Simpanan.objects.create(
#                             anggota=anggota_obj, admin=admin, jenis_simpanan=pokok_jenis,
#                             tanggal_menyimpan=tanggal_fix, jumlah_menyimpan=pokok, dana_sosial=Decimal("0"),
#                         )
#                         total_import += 1

#                     if wajib > 0:
#                         Simpanan.objects.create(
#                             anggota=anggota_obj, admin=admin, jenis_simpanan=wajib_jenis,
#                             tanggal_menyimpan=tanggal_fix, jumlah_menyimpan=wajib, dana_sosial=Decimal("0"),
#                         )
#                         total_import += 1

#                     if sukarela > 0 or dana_sosial > 0:
#                         Simpanan.objects.create(
#                             anggota=anggota_obj, admin=admin, jenis_simpanan=sukarela_jenis,
#                             tanggal_menyimpan=tanggal_fix, jumlah_menyimpan=sukarela, dana_sosial=dana_sosial,
#                         )
#                         total_import += 1

#                     start_row += 3

#             return JsonResponse({
#                 "success": True,
#                 "message": f"✅ Total {total_import} transaksi berhasil diimpor untuk {nama} ({nomor_anggota})."
#             })

#         except Exception as e:
#             import traceback
#             traceback.print_exc()
#             return JsonResponse({"success": False, "message": f"❌ Kesalahan: {str(e)}"})

#     return JsonResponse({"success": False, "message": "Tidak ada file Excel dikirim."})